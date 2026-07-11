"""
Excel Reader Module
Loads and parses Excel files from input folder.
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelReader:
    """Reads and parses Excel files."""

    def __init__(self, input_dir="input"):
        """
        Initialize ExcelReader.
        
        Args:
            input_dir: Directory containing Excel files (default: 'input')
        """
        self.input_dir = Path(input_dir)
        if not self.input_dir.exists():
            self.input_dir.mkdir(parents=True, exist_ok=True)
            logger.warning(f"Created input directory: {self.input_dir}")

    def get_excel_files(self):
        """
        Get all Excel files in input directory.
        
        Returns:
            List of Path objects for Excel files (.xlsx, .xls)
        """
        if not self.input_dir.exists():
            logger.warning(f"Input directory not found: {self.input_dir}")
            return []

        excel_extensions = ('.xlsx', '.xls')
        excel_files = [
            f for f in self.input_dir.glob('*')
            if f.is_file() and f.suffix.lower() in excel_extensions
        ]
        logger.info(f"Found {len(excel_files)} Excel files")
        return excel_files

    def load_workbook(self, file_path):
        """
        Load Excel workbook.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Workbook object or None if error
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                logger.error(f"File not found: {file_path}")
                return None

            wb = load_workbook(file_path, data_only=True)
            logger.info(f"Loaded workbook: {file_path.name}")
            return wb
        except Exception as e:
            logger.error(f"Error loading workbook {file_path}: {str(e)}")
            return None

    def get_sheet_names(self, workbook):
        """
        Get all sheet names from workbook.
        
        Args:
            workbook: Workbook object
            
        Returns:
            List of sheet names
        """
        if not workbook:
            return []
        return workbook.sheetnames

    def read_sheet_as_dict(self, workbook, sheet_name, include_header=True):
        """
        Read sheet data as list of dictionaries.
        
        Args:
            workbook: Workbook object
            sheet_name: Name of sheet to read
            include_header: If True, use first row as keys
            
        Returns:
            List of dictionaries (each row as dict) or None if error
        """
        try:
            if not workbook:
                logger.error("Invalid workbook")
                return None

            if sheet_name not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in workbook")
                return None

            ws = workbook[sheet_name]
            data = []

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                logger.warning(f"Sheet '{sheet_name}' is empty")
                return []

            if include_header and len(rows) > 1:
                headers = rows[0]
                for row in rows[1:]:
                    if any(cell is not None for cell in row):
                        row_dict = {
                            headers[i]: row[i] for i in range(len(headers))
                            if i < len(row)
                        }
                        data.append(row_dict)
            else:
                for row in rows:
                    if any(cell is not None for cell in row):
                        data.append(list(row))

            logger.info(f"Read {len(data)} rows from sheet '{sheet_name}'")
            return data

        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name}: {str(e)}")
            return None

    def read_sheet_as_list(self, workbook, sheet_name):
        """
        Read sheet data as list of lists.
        
        Args:
            workbook: Workbook object
            sheet_name: Name of sheet to read
            
        Returns:
            List of lists or None if error
        """
        try:
            if not workbook or sheet_name not in workbook.sheetnames:
                return None

            ws = workbook[sheet_name]
            data = []

            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    data.append(list(row))

            logger.info(f"Read {len(data)} rows from sheet '{sheet_name}'")
            return data

        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name}: {str(e)}")
            return None

    def get_cell_value(self, workbook, sheet_name, row, column):
        """
        Get value from specific cell.
        
        Args:
            workbook: Workbook object
            sheet_name: Sheet name
            row: Row number (1-indexed)
            column: Column letter or number
            
        Returns:
            Cell value or None
        """
        try:
            if not workbook or sheet_name not in workbook.sheetnames:
                return None

            ws = workbook[sheet_name]
            cell = ws[f"{column}{row}"] if isinstance(column, str) else ws.cell(row, column)
            return cell.value

        except Exception as e:
            logger.error(f"Error reading cell: {str(e)}")
            return None

    def read_all_files(self):
        """
        Read all Excel files in input directory.
        
        Returns:
            Dictionary with filename as key, data as value
        """
        all_data = {}
        excel_files = self.get_excel_files()

        for file_path in excel_files:
            wb = self.load_workbook(file_path)
            if wb:
                all_data[file_path.name] = {
                    'workbook': wb,
                    'sheets': {}
                }
                for sheet_name in self.get_sheet_names(wb):
                    data = self.read_sheet_as_dict(wb, sheet_name)
                    all_data[file_path.name]['sheets'][sheet_name] = data

        return all_data


def read_excel_file(file_path, sheet_name=None):
    """
    Convenience function to read a single Excel file.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Specific sheet to read (default: first sheet)
        
    Returns:
        Sheet data as list of dictionaries
    """
    reader = ExcelReader()
    wb = reader.load_workbook(file_path)
    if not wb:
        return None

    if sheet_name is None:
        sheet_name = reader.get_sheet_names(wb)[0]

    return reader.read_sheet_as_dict(wb, sheet_name)
